import os
import re
import csv
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
DOCS_DIR = r"C:\Users\PC\Documents\QNHU\US\TN\Project\GP-Travel-Advisor-Backend\docs\price_reports"
MD_PATH = os.path.join(DOCS_DIR, "missing_prices_report.md")
EXCEL_PATH = os.path.join(DOCS_DIR, "missing_prices_report.xlsx")
CSV_SLOT_TYPE_PATH = os.path.join(DOCS_DIR, "missing_prices_by_slot_type.csv")
CSV_PROVINCE_PATH = os.path.join(DOCS_DIR, "missing_prices_by_province.csv")
CSV_DETAILED_PATH = os.path.join(DOCS_DIR, "missing_prices_detailed.csv")

def parse_markdown_report(md_path):
    slot_type_data = []
    province_data = []
    detailed_data = []
    
    current_section = 0
    current_province = None
    current_group = None
    
    # Regex patterns
    slot_pattern = re.compile(r"\|\s*\*\*?(.*?)\*\*?\s*\|\s*([\d,]+)\s*\|\s*([\d,]+)\s*\|\s*([\d\.]+%)\s*\|")
    prov_pattern = re.compile(r"\|\s*([^|:]+?)\s*\|\s*([\d,]+)\s*\|\s*([\d,]+)\s*\|\s*([\d\.]+%)\s*\|")
    prov_header_pattern = re.compile(r"### 📍\s*([^(]+)")
    group_header_pattern = re.compile(r"\*\s*\*\*([^*]+)\*\*")
    detail_pattern = re.compile(r"\s*\*\s*([^:]+):\s*Thiếu\s*([\d,]+)\s*/\s*tổng\s*số\s*([\d,]+)\s*địa\s*điểm\s*\(([\d\.]+%)\)")
    
    with open(md_path, "r", encoding="utf-8") as f:
        for line in f:
            line_str = line.strip()
            
            # Detect sections
            if "## 1. Báo Cáo Tổng Quan" in line:
                current_section = 1
                continue
            elif "## 2. Thống Kê Tổng Quan Theo Tỉnh/Thành Phố" in line:
                current_section = 2
                continue
            elif "## 3. Thống Kê Chi Tiết 2 Cấp Theo Từng Tỉnh/Thành Phố" in line:
                current_section = 3
                continue
            elif "## 4." in line:
                current_section = 4
                continue
            
            # Section parsing
            if current_section == 1:
                m = slot_pattern.match(line_str)
                if m and "Nhóm dịch vụ" not in line_str and "---" not in line_str:
                    slot_type = m.group(1).replace("**", "").strip()
                    missing = int(m.group(2).replace(",", ""))
                    total = int(m.group(3).replace(",", ""))
                    pct_str = m.group(4).strip()
                    pct = float(pct_str.replace("%", "")) / 100.0
                    slot_type_data.append({
                        "Nhóm dịch vụ (slot_type)": slot_type,
                        "Số lượng thiếu giá": missing,
                        "Tổng số địa điểm": total,
                        "Tỷ lệ thiếu giá": pct
                    })
            
            elif current_section == 2:
                m = prov_pattern.match(line_str)
                if m and "Tỉnh / Thành phố" not in line_str and "---" not in line_str:
                    province = m.group(1).strip()
                    missing = int(m.group(2).replace(",", ""))
                    total = int(m.group(3).replace(",", ""))
                    pct_str = m.group(4).strip()
                    pct = float(pct_str.replace("%", "")) / 100.0
                    province_data.append({
                        "Tỉnh / Thành phố": province,
                        "Số lượng thiếu giá": missing,
                        "Tổng số địa điểm": total,
                        "Tỷ lệ thiếu giá": pct
                    })
                    
            elif current_section == 3:
                m_prov = prov_header_pattern.search(line_str)
                if m_prov:
                    current_province = m_prov.group(1).strip()
                    continue
                
                m_group = group_header_pattern.search(line_str)
                if m_group:
                    current_group = m_group.group(1).strip()
                    continue
                
                m_detail = detail_pattern.search(line_str)
                if m_detail and current_province and current_group:
                    subtype = m_detail.group(1).strip()
                    missing = int(m_detail.group(2).replace(",", ""))
                    total = int(m_detail.group(3).replace(",", ""))
                    pct_str = m_detail.group(4).strip()
                    pct = float(pct_str.replace("%", "")) / 100.0
                    detailed_data.append({
                        "Tỉnh / Thành phố": current_province,
                        "Nhóm loại hình": current_group,
                        "Loại hình chi tiết": subtype,
                        "Số lượng thiếu giá": missing,
                        "Tổng số địa điểm": total,
                        "Tỷ lệ thiếu giá": pct
                    })
                    
    return slot_type_data, province_data, detailed_data

def write_excel_premium(df_slots, df_provs, df_details, excel_path):
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_slots.to_excel(writer, sheet_name="Nhóm dịch vụ", index=False)
        df_provs.to_excel(writer, sheet_name="Tỉnh thành", index=False)
        df_details.to_excel(writer, sheet_name="Chi tiết 2 cấp", index=False)
        
        workbook = writer.book
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Dark steel blue
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Light greyish blue
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            worksheet.views.sheetView[0].showGridLines = True
            
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
            
            for row_idx in range(2, max_row + 1):
                is_zebra = (row_idx % 2 == 0)
                row_fill = zebra_fill if is_zebra else white_fill
                
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.border = thin_border
                    
                    val = cell.value
                    col_name = worksheet.cell(row=1, column=col_idx).value
                    
                    if "Tỷ lệ" in col_name or "%" in col_name:
                        cell.number_format = '0.00%'
                        cell.alignment = align_right
                    elif "Số lượng" in col_name or "Tổng số" in col_name:
                        cell.number_format = '#,##0'
                        cell.alignment = align_right
                    elif isinstance(val, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = align_right
                    else:
                        cell.alignment = align_left
            
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                col_name = worksheet.cell(row=1, column=col[0].column).value
                for cell in col:
                    if cell.value is not None:
                        if "Tỷ lệ" in col_name:
                            val_len = 8
                        elif isinstance(cell.value, float):
                            val_len = len(f"{cell.value:.2f}")
                        else:
                            val_len = len(str(cell.value))
                        max_len = max(max_len, val_len)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

def main():
    print("Parsing report from markdown...")
    slots, provs, details = parse_markdown_report(MD_PATH)
    
    df_slots = pd.DataFrame(slots)
    df_provs = pd.DataFrame(provs)
    df_details = pd.DataFrame(details)
    
    # Export to CSVs with BOM for Microsoft Excel Vietnamese language compatibility
    print("Exporting CSV files...")
    df_slots.to_csv(CSV_SLOT_TYPE_PATH, index=False, encoding="utf-8-sig")
    df_provs.to_csv(CSV_PROVINCE_PATH, index=False, encoding="utf-8-sig")
    df_details.to_csv(CSV_DETAILED_PATH, index=False, encoding="utf-8-sig")
    
    # Export to premium formatted Excel
    print("Exporting styled Excel file...")
    write_excel_premium(df_slots, df_provs, df_details, EXCEL_PATH)
    
    print("Conversion completed successfully!")
    print(f"Excel file: {EXCEL_PATH}")
    print(f"CSVs in: {DOCS_DIR}")

if __name__ == "__main__":
    main()
